import xarray as xr
import os
import os.path
import pandas as pd
import numpy as np
import re
# import matplotlib.pyplot as plt 
# # import matplotlib.ticker as mtick
# import plotly.express as px
# import plotly.graph_objects as go
import glob
import os
from itertools import combinations
from os import listdir
from os.path import join,isdir,dirname
import matplotlib.pyplot as plt
from collections import OrderedDict
from datetime import datetime
import sys
import openpyxl
from openpyxl import load_workbook,Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
#%matplotlib inline
###############################################################################
def mkdir(dir = None):
    #'''
    #Creates the given directory.

    #Parameters
    #----------
    #dir : char
    #       Directory
    #'''
    if not dir is None:
        if not os.path.exists(dir):
            os.makedirs(dir)
###############################################################################
# Function Zones
def dms2dd(degrees, minutes, seconds, direction): # Decimal minutes second to decimal degree
    dd = float(degrees) + float(minutes)/60 + float(seconds)/(60*60)
    if direction == 'W' or direction == 'S':
        dd *= -1
    return dd
###############################################################################
def dd2dms(deg):  # Decimal degree to decimal minutes second
    d = int(deg)
    md = abs(deg - d) * 60
    m = int(md)
    sd = (md - m) * 60
    return [d, m, sd]
###############################################################################
def parse_dms(dms,preserve_dd=False): # Parse string into dms
    if not preserve_dd:
        parts = re.split('[^\d\w]+', dms)
    else:
        parts = re.split('[^\d.]+', dms)
    return parts
###############################################################################
def parse_dms_from_string(string): # Parse string into dms
    # Note the pattern is pre-specified and may vary from case to case
    degree="".join(re.findall('\d+[o°]', string)).strip("o°")
    minute="".join(re.findall("[o°]\s?\d+[´]", string)).strip("o°´")
    second="".join(re.findall("[´]\d+[´´]", string)).strip("´´´")
    return [degree,minute,second]
###############################################################################
# function to remove empty rows in excel
# https://openpyxl.readthedocs.io/en/stable/styles.html#styling-merged-cells
def remove(sheet):
     # iterate the sheet by rows
    for row in sheet.iter_rows():
  
      # all() return False if all of the row value is None
        if all([cell.value is None for cell in row]):
  
      # detele the empty row
            sheet.delete_rows(row[0].row, 1)
  
      # recursively call the remove() with modified sheet data
            remove(sheet)
            
###############################################################################          
def subset_list_from_list(list_of_list,filter_list):
    '''
    Select subset of a list-of-list with selection criteria provided by filter list.

    Parameters
    ----------
    list_of_list : a list of list, user-specified input
    filter_list: a list contains elements that are contained within original sublist elements
    '''
    List_of_List_Target=[] # This is the refined list of list from input
    for sub_list in list_of_list:
        sub_list_new=[]
        for sub_list_ele in sub_list:
            if any(ele in sub_list_ele for ele in filter_list): # Check each sublist, if any elements are desired
                sub_list_new.append(sub_list_ele)
        if len(sub_list_new)!=0:
            List_of_List_Target.append(sub_list_new)
    return List_of_List_Target

###############################################################################
def extract_site_data_series(ncfile,nc_var_identifier,lon1,lat1,time_period,method="nearest"):
    '''
    Extract site-specific time-series data. Note this function only works for nc file with one variable each

    Parameters
    ----------
    ncfile : path-like object
    nc_var_identifier: str, to identify the actual variable name in nc file
    lon1: longitude in deicmal degree
    lat1: latitude in deicmal degree
    time_period: a target period to extract time-series data
    method: optional, the built-in xarray method to approximate the target grid
    '''
    ds_nc=xr.open_dataset(ncfile,mask_and_scale=True) # Always process the original data if scale_factor and add_offset exists
    # Obtain the actual coordinate/dimension names, e.g. the latitude cooridnate variable could be "latitude" or "lat" 
    # Note list(ds.dims.keys()) would yield the same results
    for dimension in list(ds_nc.coords.keys()):  
        if "lon" in dimension: # The minimal naming convenction for longitude is lon, but can be possible with full name as longitude
            lon_name="".join(re.findall(r"lon\w*",dimension)) # Get the exact longitude name 
        elif "lat" in dimension: # The minimal naming convenction for latitude is lat, but can be possible with full name as latitude
            lat_name="".join(re.findall(r"lat\w*",dimension)) # Ger the exact latitude name
    # Retrive the lat and lon vectors
    #lon_vector=ds_nc[lon_name].data
    #lat_vector=ds_nc[lat_name].data
    # Obtain the underlying variable name contained in the nc file
    var_nc_name=[item for item in list(ds_nc.data_vars) if nc_var_identifier in item][0]
    # Retrive the underly DataArray given the variable name
    var_dict=ds_nc[var_nc_name] 
    # Extract the climate data given the site coordinates and study period
    site_data=var_dict.sel({lon_name:lon1,lat_name:lat1,"time":time_period},method=method).values
    if site_data.ndim != 1:
        site_data=np.squeeze(site_data)
    # Check units and convert unit to standard one
    if var_dict.attrs["units"] in ["K","k","Kelvin","kelvin"]: 
        site_data=site_data-273.15 # Transform the kelvin to celcius
    elif var_dict.attrs['standard_name']== 'precipitation_flux' or "s-1" in var_dict.attrs['units']:
        site_data=site_data*3600*24 # Transform the precipitation flux to precipitation amount/day
    if ("rr" in nc_var_identifier) or ("pr" in nc_var_identifier):
        site_data=np.array(site_data)
        site_data[site_data<0.5]=0 # For any precipitation amount that is below 0.5 mm, set it to 0 mm.
    #### To be futher added if additional attributes area found in the nc files that are needed to transform the data
    
    # Convert the output in one-D array into pd.Series
    site_data_ser = pd.Series(site_data,index=time_period,name=nc_var_identifier)
    
    return site_data_ser
##############################################################################
def save_excel_openpy(filepath,worksheet_name,df):
    '''
    Save df to excel file using openpy method
    Parameters
    ----------
    filepath : path-like object with suffix ".xlsx"
    worksheet_name: str, specify the desired woksheet name
    df: panda dataframe
    '''
    wb = Workbook()
    #ws_active=wb.active # Create an active spreadsheet
    #ws_active.title="active sheet" 
    #ws1=wb.create_sheet(title="sheet1") # Create an excel spreadsheet
    dest_filename = join(filepath,'save.xlsx')
    ws = wb.active
    rows=dataframe_to_rows(df)
    ws.title = worksheet_name
    for r_idx, row in enumerate(rows,1):
        for c_idx, value in enumerate(row,1):
            ws.cell(row=r_idx,column=c_idx,value=value)
    wb.save(dest_filename)      
##############################################################################
def as_text(value):
    '''
    Convert any cell value into text
    Parameter
    ----------
    value : cell value
    '''
    if value is None:
        return ""
    else:
        value=str(value)
    return value
##############################################################################
def auto_fit_column_width(worksheet,reference_row,multiply_coefficient=1.5):
    '''
    Automatically adjust the column width to fit the cell content, i.e. ensure proper cell width for display
    Parameter
    ----------
    worksheet : openpyxyl worksheet object, the target worksheet to be modified with
    reference_row: int, the reference row number that is used to extract cell width for determining the entire column width 
    multiply_coefficient: float, further expand the determined column width, by default 1.5
    '''
    from openpyxl.utils import get_column_letter
    
    for i, column_cells in enumerate(worksheet.columns):
        col_letter=get_column_letter(i+1)
        length = max(len(as_text(cell.value)) for row_idx, cell in enumerate(column_cells) if row_idx==reference_row) # Check along one column the maximum width of a cell
        worksheet.column_dimensions[col_letter].width = length*multiply_coefficient #  Set the column width with the maximum width of a given cell in this column
        #worksheet.column_dimensions[col_letter].hidden = False
        #worksheet.column_dimensions[col_letter].bestFit = True
        #worksheet.column_dimensions[col_letter].auto_size = True
        #worksheet.column_dimensions[col_letter].collapsed = False
##############################################################################
def set_cell_style(worksheet,set_font_style=True,identification_character1="RCP45",identification_character2="RCP85"):
    '''
    Set the excel cell style. At present, only set cell font style is implemented and can be further complemented. 
    Also note, the iteration over the worksheet is through thr rows (equivalent of iterations over columns)
    Parameter
    ----------
    worksheet : openpyxyl worksheet object, the target worksheet as input
    identification_character1: str, character or character sequence to identify particular cell for modification
    identification_character1: str,  character or character sequence to identify particular cell for modification
    '''
    # Documentation: https://openpyxl.readthedocs.io/en/stable/styles.html  # Also check Styling Merged Cells
    from openpyxl.styles import Font
    for row_cells in worksheet.rows:
        for cell in row_cells:
            #print(cell.value)
            # Set the cell style
            if set_font_style is True:
                if identification_character1 in str(cell.value):
                    cell.font=Font(color="000000FF",bold=True) 
                elif identification_character2 in str(cell.value):
                    cell.font=Font(color="00FF0000",bold=True)
# Check Styling Merged Cells
#ws.move_range("I1:J2", rows=1)
#ws.move_range("A1:G{}".format(weather_fixed_df.shape[0]+1), rows=2)
#ws.delete_cols(weather_fixed_df.shape[1]+1)
#remove(ws) # Remove any white rows
#for col_num in np.arange(1,weather_fixed_df.shape[1]+1,1):
#    ws.merge_cells(start_row=1, start_column=col_num, end_row=2, end_column=col_num)
#ws.merge_cells(start_row=1, end_row=1,start_column=weather_fixed_df.shape[1], end_column=weather_fixed_df.shape[1]+1)
#wb.save()
#x1,y1=conver_coordinate(proj_crs_str,lon1,lat1)
###############################################################################
def consecutive_merge_df(iterable,how="outer",on=None,copy=None):
    '''
    Merge df consecutively by choosing a set of common columns (used as join keys) that are available in both df 
    
    Parameter
    ----------
    iterable : list or tuple, an iterable of df to merge. Not implemented for dict type 
    how: str, method of merging df. See panda documentations
    on: str,  a common set of columns available in all underlying df that are used as join keys
    copy: bool, if make a copy or not.
    '''
    # Check if the input typ is a list or tuple
    if not isinstance(iterable,(list,tuple)):
        raise TypeError("The input iterable is not a required data type, only list or tuple is supported, but {} is found".format(type(iterable))) 
    for i in range(len(iterable)):
        if i != (len(iterable)-1):
            iterable[i+1] = iterable[i].merge(iterable[i+1], how=how, on=on, copy=copy) # Update the next df in place from the input iterable
        else:
            continue
    # Obtain the final df 
    final_df = iterable[-1] # With this method, only the last element from the iterable is required.
    return final_df
###############################################################################
# 1. Specify all user-dependent path 
root_path=r"H:\Grapevine_model_VineyardR_packages_SCE_implementations"
phenology_path = join(root_path,"phenology_data")
weather_station_path= join(root_path,"weather_station_data")
gridded_IP_path = join(root_path,"gridded_climate_data")
meta_data_path = join(root_path,"metadata")
###############################################################################
# 2. Load and process the phenology observational file
pheno_datasets = glob.glob(join(phenology_path,"*UTAD*.xlsx"))
# 2.1 Check if all location names are consistent among different phenology datasets
location_full = []
for pheno_file in pheno_datasets:
    load_file = pd.ExcelFile(pheno_file)
    location_full.append(load_file.sheet_names)
# Iterate over the appended lists to see if list names are identical
for location_list in combinations(location_full,2): # Return r length subsequences of elements from the input iterable
    if not location_list[0].__eq__(location_list[-1]):
        print("Location names are inconsistent in different dataset files")
        break
    else:
        pass
# 2.2 Process the phenology data into the desired format
target_varieties = ["Touriga Francesa","Touriga Nacional"]
# Define a list of early varieties
early_varieties = ["Touriga Francesa", "Jaen", "Aragonez", "Tinta Barroca", "Tinta Francisca", "Fernão Pires", "Alfrocheiro"]
#intermediate_varieties
# Define a list of late varieties
late_varieties = ["Avesso", "Azal", "Arinto", "Encruzado", "Espadeiro", "Viosinho", "Tinto Cão"]

full_BBCH = []
for BBCH_data in pheno_datasets: # Iterate over the searched glob list for the phenology datasets
    # Extract the phenology variable name
    BBCH_name = [name for name in ["budburst","flowering","veraison"] if name in BBCH_data][0]
    location_list = np.unique(location_full) # Obtain the unique list of location in the dataset
    target_df=pd.DataFrame()
    for location_name in location_list:
        location_df = pd.read_excel(BBCH_data,sheet_name = location_name)
        if not location_df.empty: # If not empty
            for row_data in location_df.itertuples(index=False,name=None):
                for row_ele in row_data:
                    if any(variety in str(row_ele) for variety in early_varieties):
                        target_col_name = location_df.columns[row_data.index(row_ele)] # Get the target column name
                        target_data_ser = location_df.loc[location_df[target_col_name]==row_ele,:] # Load the time series of phenolgoy data given the target location name 
                        target_col = [ele for ele in target_data_ser.columns if re.search(r"\d+",str(ele))] # Look for data columns only
                        target_data_ser = target_data_ser.loc[:,target_col] # Select the data series
                        target_data_ser.dropna(axis=1, how='any',inplace=True) # Drop Na values
                        target_data_ser = target_data_ser.T # Transpose the df
                        # Create the target df given a local site and variety 
                        target_data_df = pd.DataFrame({"Sites":location_name,"Years":target_data_ser.index,
                                                       "Varieties":row_ele,BBCH_name+"(DOY)":np.array(target_data_ser).squeeze()})
                        # Append the data df to the target data df
                        target_df = pd.concat([target_df,target_data_df], axis=0, join='outer', ignore_index=True)
                        #target_df.append(target_data_df,ignore_index=True,verify_integrity=True)
                    else:
                        continue
    # Append the target_df to an existing empty list
    full_BBCH.append(target_df)
# 2.3. Write the final merged df into excel file
# 2.3.1 Obtain the final merged df
merged_df = consecutive_merge_df(full_BBCH,how="outer",on=["Sites","Years","Varieties"],copy=True)
merged_df.fillna(-999, axis=0, inplace=True)
#Lisbon_data = merged_df.loc[merged_df["Sites"]=="Lisboa", :]
# 2.3.2 Write the merged df into excel file
save_path_pheno = join(phenology_path,"summary")
mkdir(save_path_pheno)
#writer = pd.ExcelWriter(join(save_path,"summary_phenology.xlsx")
                      
# 2.3.3 Save the df into .excel or .csv file in the disk
merged_df.to_csv(join(save_path_pheno,"summary_phenology.csv"), sep=",",header=True,
                        index=True,encoding ="utf-8") # Save to .excel 

# Lisbon_data.to_csv(join(save_path_pheno,"summary_phenology_lisboa.csv"), sep=",",header=True,
#                         index=True,encoding ="utf-8") # Save to .excel 

# merged_df.to_excel(writer,sheet_name="summary_phenology",header=True,
#                               index=True,engine="openpyxl") # Save to .csv
# writer.save() # Save to disk
###############################################################################
# 3. Extract weather data for each station
# 3.1 Create correct weather variable names for the file
weather_var_dict= {"Tn":"tasmin", "Tg":"tas",
          "Tx":"tasmax"
          }
# 3.2 Desired order of column for the output weather file
weather_col_order = ['day','month','Year', 'tasmax', 'tas', 'tasmin', 'Lat','Lon']
grid_weather_col_order = ['day','month','Year', 'tn', 'tg', 'tx', "rr", 'Lat','Lon']
# 3.3 Original weather station file
weather_station_data_path = glob.glob(join(weather_station_path,"*.xlsx"))[0]
# 3.4 Inquire the list of stations in the data file
weather_station_load = pd.ExcelFile(weather_station_data_path)
save_path_weather = join(weather_station_path,"summary_weather")
mkdir(save_path_weather)
# 3.5 Define the path to the Iberian datasets
iberian_data_path = r"H:\E_OBS_V24"
var_identifiers = ["rr", "tg", "tn","tx"] # ['tasmax', 'tas', 'tasmin',"pr"]
target_dataset = [] # Collect target datasets
for ncfile in glob.glob(join(iberian_data_path,"*.nc")):
    if any(varname in ncfile for varname in var_identifiers):
        target_dataset.append(ncfile)
mkdir(gridded_IP_path)
# 3.6 Iterate over each station
for station_name in weather_station_load.sheet_names:
    weather_station_data = pd.read_excel(weather_station_data_path,sheet_name=station_name)
    weather_station_data.rename(columns=weather_var_dict,inplace=True)
    # Create an intermediate string series in the format of YEAR-DOY
    datetime_str =  weather_station_data["Year"].astype(str)+ "-" +weather_station_data["DOY"].astype(str)
    #weather_station_data.apply(lambda x: datetime.strptime('{}-{}'.format(x["Year"].astype(str), x["DOY"].astype(str)),'%Y-%j')
    # Create two additional columns for day and months
    weather_station_data["day"] = datetime_str.apply(lambda x: datetime.strptime(x,'%Y-%j').day)
    weather_station_data["month"] = datetime_str.apply(lambda x: datetime.strptime(x,'%Y-%j').month)
    # Reorder the df
    weather_station_data = weather_station_data.reindex(columns=weather_col_order,copy=True)
    # Check for missing values
    if weather_station_data.isnull().values.any():
        weather_station_data.fillna(method="ffill",inplace=True)
    # Save the csv file into disk for the location
    #weather_station_data.to_csv(join(save_path_weather,"{}_no_NAN.csv".format(station_name)), sep=",",header=True,
                            #index=False,encoding ="utf-8")
    # Extract the coordinate that is going to be used to extract the weather data for Iberian peninsula or E-OBS gridded dataset
    # Look for the lon and lat pair
    lon= np.unique([round(ele,3) for ele in weather_station_data["Lon"].unique()])
    lat= np.unique([round(ele,3) for ele in weather_station_data["Lat"].unique()])
    if len(lon)>1 or len(lat)>1:
        print("The geographic coordinate lat and lon are not unique")
        break
    else:
        lon = float(lon)
        lat = float(lat)
    target_data_list =[]
    for grid_data in target_dataset:
        time_period = xr.open_dataset(grid_data,mask_and_scale=True).time.data # Extract the full time period
        var_shortname = [var for var in var_identifiers if var+"_" in grid_data][0] # Extract the variable short name
        data_series = extract_site_data_series(grid_data,var_shortname,lon,lat,time_period) # Extract the time series of data for a given site
        target_data_list.append(data_series) # Append to an existing empty list
    # Concatenate all list datasets 
    merged_df = pd.concat(target_data_list,axis=1, join='inner', ignore_index=False)
    # Create a df with all desired columns
    merged_df = merged_df.assign(day=lambda x:x.index.day, month=lambda x:x.index.month, Year= lambda x:x.index.year,
                     Lon=lon, Lat=lat)
    # Reorder the columns
    merged_df_final = merged_df.reindex(columns = grid_weather_col_order, copy=True)
    # Check for missing values
    if merged_df_final.isnull().values.any():
        merged_df_final.fillna(method="ffill",inplace=True)
    # Export the target df into csv file
    merged_df_final.to_csv(join(gridded_IP_path,"{}_gridded_data.csv".format(station_name)), sep=",",header=True,
                            index=False,encoding ="utf-8")
###############################################################################

# Define the time range
start_year = 2017
end_year = 2018
# Obtain the list of files for 2020 and 2021
#E_OB_nc_2020=[ncfile for ncfile in glob.glob(join(E_OBS_path,"*.nc")) if ("rr" in ncfile) or ("tg" in ncfile)]
#E_OB_nc_2021=[ncfile for ncfile in glob.glob(join(dirname(Root_path),"E_OBS_2021","*.nc")) if ("rr" in ncfile) or ("tg" in ncfile)]
## Define the study period
#year_2020=pd.date_range(str(start_year)+"-01-01",str(start_year)+"-12-31", freq="D")
#year_2021=pd.date_range(str(start_year)+"-01-01",str(end_year)+"-8-31", freq="D")
time_series=pd.date_range(str(start_year)+"-01-01",str(end_year)+"-12-31", freq="D")
#metadata_df=pd.DataFrame({"time":time_series,"Lat":lat1,"Lon":lon1,"Year":time_series.year,
                                     # "Month":time_series.month,"Day_in_month":time_series.day,
                                        #"DOY":time_series.day_of_year}).set_index("time")
# Extraction data for the mean temperature
var_dict= {"tn":"Minimum daily surface 2-m air temperature (°C)", "tg":"Mean daily surface 2-m air temperature (°C)",
          "tx":"Maximum daily surface 2-m air temperature (°C)", "rr":"Daily surface 2-m precipitation sum (mm)",
          "qq":"Mean daily global radiation (w/m2)", "fg":"Mean daily surface 2-m wind speed (m/s)", "hu":"Mean daily surface 2-m relative humidity (%)"
          }


data_ser_list=[]
data_ser_resample_list=[]
for var_name in var_dict.keys():
    var_fullname = var_dict[var_name]
    # Get the nc file in 2020 and 2021
    nc_file = [file for file in glob.glob(join(E_OBS_path,"*.nc")) if var_name in file][0]
    #nc_2020=[file for file in E_OB_nc_2020 if var in file][0]
    #nc_2021=[file for file in E_OB_nc_2021 if var in file][0]
    # Extract the data for the given site in 2020 and 2021
    site_data =extract_site_data_series(nc_file,var_name,lon1=lon1,lat1=lat1,time_period=time_series)
    site_data.rename(var_fullname,inplace=True)
    # Append panda series data into the list
    data_ser_list.append(site_data)# Resample the timeseries data
    if "rr" not in var_name: # For temperature variables 
        site_data_resample= site_data.resample("M").mean()
    else: # For precipitation variables
        site_data_resample= site_data.resample("M").sum()
    #data_2021=extract_site_data_series(nc_2021,var,lon1=lon1,lat1=lat1,time_period=year_2021)
    # Concatenate the series data from two years
    #data_concat=np.concatenate([data_2020,data_2021],axis=0)
    # Append the resampled timeseries data into the list
    data_ser_resample_list.append(site_data_resample)
    
# Obtain the concatenated df between the two variables
data_output_df=pd.concat(data_ser_list,axis=1,join="inner")
data_output_df_resample=pd.concat(data_ser_resample_list,axis=1,join="inner")
# Concatenate the data columns with the label columns
target_df = pd.concat([metadata_df,data_output_df],axis=1,join='inner')
target_df_resample = pd.concat([metadata_df,data_output_df_resample],axis=1,join='inner')
target_df_resample.drop(columns=['Day_in_month', 'DOY'],inplace=True)
# Write the df to file
writer=pd.ExcelWriter(join(Root_path,"Luis_DATA_{}_{}.xlsx".format(str(start_year),str(end_year))))
target_df.to_excel(writer,sheet_name="daily_data",header=True,
                              index=True,engine="openpyxl")
target_df_resample.to_excel(writer,sheet_name="monthly_data",header=True,
                              index=True,engine="openpyxl")
writer.save()
# Read the processed ET0 data
# ws_data = [ file for file in glob.glob(join(Root_path,"*.xlsx")) if "2017_2018" in file][0]
# excel_data= pd.read_excel(ws_data,sheet_name="daily_data")
# excel_data.rename(columns={"Unnamed: 0":"timeindex"},inplace=True)
# excel_data.set_index("timeindex",inplace=True)
# et0_col_name = [name for name in excel_data.columns if "ET0" in name][0]
# et0_ser = excel_data[et0_col_name]
# et0_ser_monthly = et0_ser.resample("M").sum()
###############################################################################
###############################################################################

###############################################################################
# 2. Pre-process the metadata excel file where the coordinates are transformed into decimal degrees
Root_path = r"H:\Grapevine_vineyard_model_R_LIST"
Meta_file=glob.glob(join(Root_path,"*Meta*"))[0]
Meta_file_excel=pd.ExcelFile(Meta_file)
Meta_df=pd.read_excel(Meta_file_excel,sheet_name=[ele for ele in Meta_file_excel.sheet_names if "yield" in ele][0])
#Location_list=[]
Lat_dict={}
Lon_dict={}
for row in Meta_df.itertuples(index=False):
    ## Obtain the string format of geographic lat and lon
    site_name=row[0] # The first column for the site name
    lat=row[1] # The second column for lat
    lon=row[2] # The third column for lon
    ## Convert the lat and lon into decimal degree
    # Obtain the string format of dms for lat
    dms_string_lat=parse_dms_from_string(lat) # Return a list of dms string
    dd_lat=dms2dd(dms_string_lat[0],dms_string_lat[1],dms_string_lat[2],"N") # Obtain the latitude of the grid point
    # Obtain the string format of dms for long
    dms_string_lon=parse_dms_from_string(lon) # Return a list of dms string
    dd_lon=dms2dd(dms_string_lon[0],dms_string_lon[1],dms_string_lon[2],"W") # Obtain the latitude of the grid point
    # Update processed data into dict
    Lat_dict[site_name]=round(dd_lat,3)
    Lon_dict[site_name]=round(dd_lon,3)
# Lon_ser = Meta_df["LON"]
# Lat_ser = Meta_df["LAT"]
# # Calculate the float value for lon
# Lon_ser_str = Lon_ser.apply(lambda x: parse_dms(x,preserve_dd=True))
# Lon_ser_str = Lon_ser_str.apply(lambda x: [ele_str for ele_str in x if ele_str!=""])
# Lon_ser_float = Lon_ser_str.apply(lambda x: dms2dd(x[0],x[1],x[2],"W"))
# # Calculate the float value for lat
# Lat_ser_str = Lat_ser.apply(lambda x: parse_dms(x,preserve_dd=True))
# Lat_ser_str = Lat_ser_str.apply(lambda x: [ele_str for ele_str in x if ele_str!=""])
# Lat_ser_float = Lat_ser_str.apply(lambda x: dms2dd(x[0],x[1],x[2],"N"))
# # Create a lon, lat dataframe to export
# lon_lat_df = pd.DataFrame({"lon":Lon_ser_float,"lat":Lat_ser_float})

# Create the new metadata df
Meta_df_new=pd.DataFrame({"sites":list(Lat_dict.keys()),"lon":list(Lon_dict.values()),
                          "lat":list(Lat_dict.values())})
# Export the processed df into a csv file
Meta_df_new.to_excel(join(Root_path,"Metadata_sites_final.xlsx"),sheet_name="Sites_metadata")
##########################################################################################################################
########################################################################################################################## 
## 3. Extract the site-specific climate data
# Define the target study periods
Study_periods={"Baseline_OB_EOBS":np.arange(1991,2021,1),"Baseline_SM_ControlRun":np.arange(1991,2021,1),
               "Future_2041_2070":np.arange(2041,2071,1),"Future_2071_2100":np.arange(2071,2101,1)}
# Note the bias-adjsuted period for RCM is 1989–2010 using observational datasets from MESAN. See reference: Yang et al. 2019;
study_var=["tn","tx","rr"]
Key_order=["tg","tn","tx","rr"]
CF_name={key:value for key,value in zip(study_var,["tasmin","tasmax","pr"])} # 
output_path=join(Root_path,"save_path")
mkdir(output_path)
# Iterate over each site
for row_values in Meta_df_new.itertuples(index=False):
    site=row_values[0]
    lon1=row_values[1]
    lat1=row_values[2]
    Target_excel=join(output_path,str(site)+".xlsx")
    writer=pd.ExcelWriter(Target_excel,engine='openpyxl')
    ## Test Session ##
    # For each site, iterate over each extraction period, which will be the 
    for study_period_name, study_period in Study_periods.items():
        # Obtain the fixed column information based on the study period
        start_year=int(min(study_period))
        end_year=int(max(study_period))
        time_period=pd.date_range(str(start_year)+"-01-01",str(end_year)+"-12-31", freq="D")
        weather_fixed_columns={"Site_name":site,"Lat":lat1,"Lon":lon1,"Year":time_period.year,
                                  "Month":time_period.month,"Day_in_month":time_period.day,
                                  "DOY":time_period.day_of_year}
        # Make the fixed metadata columns that will be then merged with data columns
        weather_fixed_df=pd.DataFrame(weather_fixed_columns) # Initial the df with basic information
        if "OB" in study_period_name: # For baseline period with E-OBS observatoinal datasets
            nc_file_list=E_OB_nc.copy()
            extract_variables=["tg"]+study_var
            site_data_dict=OrderedDict()
            for ncfile in nc_file_list:
                nc_var_identifier=[ nc_var for nc_var in extract_variables if nc_var in ncfile][0]
                site_data=extract_site_data_series(ncfile,nc_var_identifier,lon1,lat1,time_period)
                var_unit_eobs="(°C)" if nc_var_identifier in ["tg","tn","tx"] else "(mm)" 
                site_data_dict[nc_var_identifier+" "+ var_unit_eobs]=site_data
            # Order the dictionary accoridng to a pre-defined order
            for key_fixed in Key_order:
                key_dict=list(site_data_dict.keys())
                order_key=[key_update for key_update in key_dict if key_fixed in key_update][0]
                site_data_dict.move_to_end(order_key,last=True)            
            # Create the df from collected dict
            site_data_dict_df=pd.DataFrame(site_data_dict)
            target_df=pd.concat([weather_fixed_df,site_data_dict_df],axis=1,join='inner')
        else:
            nc_file_list=RCM_files.copy() # RCM files contain both RCP4.5 and RCP8.5 nc files
            extract_variables=study_var.copy()
            site_data_dict=OrderedDict()
            for extract_var in extract_variables:               
                if not extract_var in site_data_dict.keys():
                    site_data_dict[extract_var]={}
                for ncfiles in nc_file_list:
                    ncfile_path=[item for item in ncfiles if extract_var+".nc" in item][0]
                    nc_var_identifier= CF_name[extract_var]
                    site_data=extract_site_data_series(ncfile_path,nc_var_identifier,lon1,lat1,time_period)
                    rcm_name="".join(re.findall(r"[\\]\w+[-]\w+[(]",ncfile_path)).strip("\(") # Extract the RCM name given the file path
                    scenario_name="".join(re.findall(r"[\\]RCP\d+[\\]",ncfile_path)).strip("\\")
                    site_data_dict[extract_var][rcm_name+"-"+scenario_name]=np.array(site_data)
            # Create an additional variable for Tmean
            T_min_dict=site_data_dict["tn"]
            T_max_dict=site_data_dict["tx"]
            T_mean_key="tg"
            for (RCM_tmin_name,RCM_tmin), (RCM_tmax_name,RCM_tmax) in zip(T_min_dict.items(),T_max_dict.items()):
                if T_mean_key not in site_data_dict.keys(): # Create a new nested dictioanry variable
                    site_data_dict[T_mean_key]={}
                if RCM_tmin_name==RCM_tmax_name: # If both key name are the same, adopt the first one as the key
                    T_mean=(RCM_tmin+RCM_tmax)/2 # Compute the mean value
                    site_data_dict[T_mean_key][RCM_tmin_name]=T_mean
                else:
                    raise ValueError("The key values from Tmin_dict and Tmax_dict are inconsistent")
                    break
            # Order the dictionary accoridng to a pre-defined order
            for key_fixed_rcm in Key_order:
                key_dict_rcm=list(site_data_dict.keys())
                order_key=[key_update for key_update in key_dict_rcm if key_fixed_rcm in key_update][0]
                site_data_dict.move_to_end(order_key,last=True)  
               
                #data_columns_df=data_columns_df.reset_index()
                # if iter_number==0: # In the first iteration over studied variables, create an empty df first while in the following iterations, the df is updated automatically
                #     target_df=pd.DataFrame()
                #     #target_df=target_df.merge(data_columns_df, how='inner',left_index=True,right_index=True)
                #     target_df[]
                # else:
                #     target_df=target_df.merge(data_columns_df, how='inner',left_index=True,right_index=True)
            # A new dictionary to store df
            site_data_dict_new={}
            for var_target,dict_var in site_data_dict.items():
                data_columns_df=pd.DataFrame(dict_var)
                var_unit="(°C)" if var_target in ["tg","tn","tx"] else "(mm)" # Specify the variable unit
                data_columns_df.columns = pd.MultiIndex.from_product([[var_target+" "+var_unit], data_columns_df.columns])
                site_data_dict_new[var_target]=data_columns_df
            # Construct the target df by concatenating all df
            concat_site_data_df=pd.concat(list(site_data_dict_new.values()),axis=1,join='inner',ignore_index=False)
            # Make the weather fixed columns with multi-level index
            weather_fixed_df.columns=pd.MultiIndex.from_product([["Station_metadata"], weather_fixed_df.columns])
            # Concate the data df with metadata df
            target_df=pd.concat([weather_fixed_df,concat_site_data_df],axis=1,join='inner')
        # Save all dfs to respective worksheet within an existing excel file
        target_df.to_excel(writer,sheet_name=study_period_name,header=True,
                              index=True,engine="openpyxl",merge_cells=True)
    # Save to excel file only after export data to all relevant worksheets
    writer.save()
    # Imediately after saving, modify the excel file to format necessary columns
    wb=load_workbook(Target_excel)
    for sheet_name in wb.sheetnames:
        ws=wb[sheet_name] # Iterate over each worksheet
        # Modify the excel worksheet format and display properties
        if "OB" in sheet_name:
            auto_fit_column_width(ws,0) # Auto-fit the column width based the first row´s cell width
            continue
        else:
            auto_fit_column_width(ws,1) # Auto-fit the column width based on the second row´s cell width
            set_cell_style(ws,identification_character1="RCP45",identification_character2="RCP85") # Set the cell colors for climate data extracted from RCMs under RCP45 and RCP85
    wb.save(Target_excel) # save the formatted excel file
##########################################################################################################################
########################################################################################################################## 
## 4. A quick check if extracted data at 2 different times are equal
from os import listdir
from os.path import join,isdir,isfile
import pandas as pd


Extract1=join(output_path,"extract1")
Extract2=output_path
# List all excel files within each extraction folder
Extract1_excels=[ excel_file for excel_file in listdir(Extract1) if isfile(join(Extract1,excel_file))] 
Extract2_excels=[ excel_file for excel_file in listdir(Extract2) if isfile(join(Extract2,excel_file))] 
# Dump excel files with the same file name from two different folders into a variable
Excel_file_names=[ excel1 for excel1, excel2 in zip(Extract1_excels,Extract2_excels) if excel1==excel2]
# Loop through these excel files to check if the data in 2 excel files are identical
for file_name in Excel_file_names:
    # Define the path to the files
    extract1_excel=join(Extract1,file_name)
    extract2_excel=join(Extract2,file_name)
    # Load the excel files into memory in order to inquire the sheet name information
    extrac1_excel_load=pd.ExcelFile(extract1_excel)
    extrac2_excel_load=pd.ExcelFile(extract2_excel)
    # Read available excel sheet names and check if the sheet_names are identical
    sheet_names=[sheet1 for sheet1, sheet2 in zip(extrac1_excel_load.sheet_names,extrac2_excel_load.sheet_names) if sheet1==sheet2]
    for sheet_name in sheet_names:
        # Read each excel sheet in to dataframe
        extract1_df=pd.read_excel(extract1_excel,sheet_name=sheet_name)
        extract2_df=pd.read_excel(extract2_excel,sheet_name=sheet_name)
        # Test if two dataframes resulting from two excel worksheets are identical
        if extract1_df.equals(extract2_df):
            print("The data for site {0} over {1} are identical".format(file_name,sheet_name))
        else:
            raise ValueError("The data for site {0} over {1} are inconsistent, errors occur for one of the file".format(file_name,sheet_name))
            break

    
    
    