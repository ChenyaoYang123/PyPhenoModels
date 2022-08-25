import xarray as xr
import os
import os.path
import pandas as pd
import numpy as np
import re
import getpass
# import matplotlib.pyplot as plt 
# # import matplotlib.ticker as mtick
# import plotly.express as px
# import plotly.graph_objects as go
import glob
import os
from itertools import combinations
from os import listdir
from os.path import join,isdir,dirname
import matplotlib.pyplot as plt
from collections import OrderedDict
from datetime import datetime
import sys
import openpyxl
from openpyxl import load_workbook,Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
#%matplotlib inline
###############################################################################
def mkdir(dir = None):
    #'''
    #Creates the given directory.

    #Parameters
    #----------
    #dir : char
    #       Directory
    #'''
    if not dir is None:
        if not os.path.exists(dir):
            os.makedirs(dir)
###############################################################################
# Function Zones
def dms2dd(degrees, minutes, seconds, direction): # Decimal minutes second to decimal degree
    dd = float(degrees) + float(minutes)/60 + float(seconds)/(60*60)
    if direction == 'W' or direction == 'S':
        dd *= -1
    return dd
###############################################################################
def dd2dms(deg):  # Decimal degree to decimal minutes second
    d = int(deg)
    md = abs(deg - d) * 60
    m = int(md)
    sd = (md - m) * 60
    return [d, m, sd]
###############################################################################
def parse_dms(dms,preserve_dd=False): # Parse string into dms
    if not preserve_dd:
        parts = re.split('[^\d\w]+', dms)
    else:
        parts = re.split('[^\d.]+', dms)
    return parts
###############################################################################
def parse_dms_from_string(string): # Parse string into dms
    # Note the pattern is pre-specified and may vary from case to case
    degree="".join(re.findall('\d+[o°]', string)).strip("o°")
    minute="".join(re.findall("[o°]\s?\d+[´]", string)).strip("o°´")
    second="".join(re.findall("[´]\d+[´´]", string)).strip("´´´")
    return [degree,minute,second]
###############################################################################
# function to remove empty rows in excel
# https://openpyxl.readthedocs.io/en/stable/styles.html#styling-merged-cells
def remove(sheet):
     # iterate the sheet by rows
    for row in sheet.iter_rows():
  
      # all() return False if all of the row value is None
        if all([cell.value is None for cell in row]):
  
      # detele the empty row
            sheet.delete_rows(row[0].row, 1)
  
      # recursively call the remove() with modified sheet data
            remove(sheet)
            
###############################################################################          
def subset_list_from_list(list_of_list,filter_list):
    '''
    Select subset of a list-of-list with selection criteria provided by filter list.

    Parameters
    ----------
    list_of_list : a list of list, user-specified input
    filter_list: a list contains elements that are contained within original sublist elements
    '''
    List_of_List_Target=[] # This is the refined list of list from input
    for sub_list in list_of_list:
        sub_list_new=[]
        for sub_list_ele in sub_list:
            if any(ele in sub_list_ele for ele in filter_list): # Check each sublist, if any elements are desired
                sub_list_new.append(sub_list_ele)
        if len(sub_list_new)!=0:
            List_of_List_Target.append(sub_list_new)
    return List_of_List_Target

###############################################################################
def extract_site_data_series(ncfile,nc_var_identifier,lon1,lat1,time_period,method="nearest"):
    '''
    Extract site-specific time-series data. Note this function only works for nc file with one variable each

    Parameters
    ----------
    ncfile : path-like object
    nc_var_identifier: str, to identify the actual variable name in nc file
    lon1: longitude in deicmal degree
    lat1: latitude in deicmal degree
    time_period: a target period to extract time-series data
    method: optional, the built-in xarray method to approximate the target grid
    '''
    ds_nc=xr.open_dataset(ncfile,mask_and_scale=True) # Always process the original data if scale_factor and add_offset exists
    # Obtain the actual coordinate/dimension names, e.g. the latitude cooridnate variable could be "latitude" or "lat" 
    # Note list(ds.dims.keys()) would yield the same results
    for dimension in list(ds_nc.coords.keys()):  
        if "lon" in dimension: # The minimal naming convenction for longitude is lon, but can be possible with full name as longitude
            lon_name="".join(re.findall(r"lon\w*",dimension)) # Get the exact longitude name 
        elif "lat" in dimension: # The minimal naming convenction for latitude is lat, but can be possible with full name as latitude
            lat_name="".join(re.findall(r"lat\w*",dimension)) # Ger the exact latitude name
    # Retrive the lat and lon vectors
    #lon_vector=ds_nc[lon_name].data
    #lat_vector=ds_nc[lat_name].data
    # Obtain the underlying variable name contained in the nc file
    var_nc_name=[item for item in list(ds_nc.data_vars) if nc_var_identifier in item][0]
    # Retrive the underly DataArray given the variable name
    var_dict=ds_nc[var_nc_name] 
    # Extract the climate data given the site coordinates and study period
    site_data=var_dict.sel({lon_name:lon1,lat_name:lat1,"time":time_period},method=method).values
    if site_data.ndim != 1:
        site_data=np.squeeze(site_data)
    # Check units and convert unit to standard one
    if var_dict.attrs["units"] in ["K","k","Kelvin","kelvin"]: 
        site_data=site_data-273.15 # Transform the kelvin to celcius
    elif var_dict.attrs['standard_name']== 'precipitation_flux' or "s-1" in var_dict.attrs['units']:
        site_data=site_data*3600*24 # Transform the precipitation flux to precipitation amount/day
    if ("rr" in nc_var_identifier) or ("pr" in nc_var_identifier):
        site_data=np.array(site_data)
        site_data[site_data<0.5]=0 # For any precipitation amount that is below 0.5 mm, set it to 0 mm.
    #### To be futher added if additional attributes area found in the nc files that are needed to transform the data
    
    # Convert the output in one-D array into pd.Series
    site_data_ser = pd.Series(site_data,index=time_period,name=nc_var_identifier)
    
    return site_data_ser
##############################################################################
def save_excel_openpy(filepath,worksheet_name,df):
    '''
    Save df to excel file using openpy method
    Parameters
    ----------
    filepath : path-like object with suffix ".xlsx"
    worksheet_name: str, specify the desired woksheet name
    df: panda dataframe
    '''
    wb = Workbook()
    #ws_active=wb.active # Create an active spreadsheet
    #ws_active.title="active sheet" 
    #ws1=wb.create_sheet(title="sheet1") # Create an excel spreadsheet
    dest_filename = join(filepath,'save.xlsx')
    ws = wb.active
    rows=dataframe_to_rows(df)
    ws.title = worksheet_name
    for r_idx, row in enumerate(rows,1):
        for c_idx, value in enumerate(row,1):
            ws.cell(row=r_idx,column=c_idx,value=value)
    wb.save(dest_filename)      
##############################################################################
def as_text(value):
    '''
    Convert any cell value into text
    Parameter
    ----------
    value : cell value
    '''
    if value is None:
        return ""
    else:
        value=str(value)
    return value
##############################################################################
def auto_fit_column_width(worksheet,reference_row,multiply_coefficient=1.5):
    '''
    Automatically adjust the column width to fit the cell content, i.e. ensure proper cell width for display
    Parameter
    ----------
    worksheet : openpyxyl worksheet object, the target worksheet to be modified with
    reference_row: int, the reference row number that is used to extract cell width for determining the entire column width 
    multiply_coefficient: float, further expand the determined column width, by default 1.5
    '''
    from openpyxl.utils import get_column_letter
    
    for i, column_cells in enumerate(worksheet.columns):
        col_letter=get_column_letter(i+1)
        length = max(len(as_text(cell.value)) for row_idx, cell in enumerate(column_cells) if row_idx==reference_row) # Check along one column the maximum width of a cell
        worksheet.column_dimensions[col_letter].width = length*multiply_coefficient #  Set the column width with the maximum width of a given cell in this column
        #worksheet.column_dimensions[col_letter].hidden = False
        #worksheet.column_dimensions[col_letter].bestFit = True
        #worksheet.column_dimensions[col_letter].auto_size = True
        #worksheet.column_dimensions[col_letter].collapsed = False
##############################################################################
def set_cell_style(worksheet,set_font_style=True,identification_character1="RCP45",identification_character2="RCP85"):
    '''
    Set the excel cell style. At present, only set cell font style is implemented and can be further complemented. 
    Also note, the iteration over the worksheet is through thr rows (equivalent of iterations over columns)
    Parameter
    ----------
    worksheet : openpyxyl worksheet object, the target worksheet as input
    identification_character1: str, character or character sequence to identify particular cell for modification
    identification_character1: str,  character or character sequence to identify particular cell for modification
    '''
    # Documentation: https://openpyxl.readthedocs.io/en/stable/styles.html  # Also check Styling Merged Cells
    from openpyxl.styles import Font
    for row_cells in worksheet.rows:
        for cell in row_cells:
            #print(cell.value)
            # Set the cell style
            if set_font_style is True:
                if identification_character1 in str(cell.value):
                    cell.font=Font(color="000000FF",bold=True) 
                elif identification_character2 in str(cell.value):
                    cell.font=Font(color="00FF0000",bold=True)
# Check Styling Merged Cells
#ws.move_range("I1:J2", rows=1)
#ws.move_range("A1:G{}".format(weather_fixed_df.shape[0]+1), rows=2)
#ws.delete_cols(weather_fixed_df.shape[1]+1)
#remove(ws) # Remove any white rows
#for col_num in np.arange(1,weather_fixed_df.shape[1]+1,1):
#    ws.merge_cells(start_row=1, start_column=col_num, end_row=2, end_column=col_num)
#ws.merge_cells(start_row=1, end_row=1,start_column=weather_fixed_df.shape[1], end_column=weather_fixed_df.shape[1]+1)
#wb.save()
#x1,y1=conver_coordinate(proj_crs_str,lon1,lat1)
###############################################################################
def consecutive_merge_df(iterable,how="outer",on=None,copy=None):
    '''
    Merge df consecutively by choosing a set of common columns (used as join keys) that are available in both df 
    
    Parameter
    ----------
    iterable : list or tuple, an iterable of df to merge. Not implemented for dict type 
    how: str, method of merging df. See panda documentations
    on: str,  a common set of columns available in all underlying df that are used as join keys
    copy: bool, if make a copy or not.
    '''
    # Check if the input typ is a list or tuple
    if not isinstance(iterable,(list,tuple)):
        raise TypeError("The input iterable is not a required data type, only list or tuple is supported, but {} is found".format(type(iterable))) 
    for i in range(len(iterable)):
        if i != (len(iterable)-1):
            iterable[i+1] = iterable[i].merge(iterable[i+1], how=how, on=on, copy=copy) # Update the next df in place from the input iterable
        else:
            continue
    # Obtain the final df 
    final_df = iterable[-1] # With this method, only the last element from the iterable is required.
    return final_df
###############################################################################
# 1. Specify all user-dependent path 
if getpass.getuser() == 'Clim4Vitis':
    script_drive = "H:\\"
elif getpass.getuser() == 'Admin':
    script_drive = "G:\\"

root_path= join(script_drive, "Grapevine_model_VineyardR_packages_SCE_implementations")
phenology_path = join(root_path,"observed_phenology")
#weather_station_path= join(root_path,"weather_station_data")
gridded_climate_path = join(root_path,"gridded_climate_data")
#meta_data_path = join(root_path,"metadata")
###############################################################################
# 2. Extract weather data for each station
# 2.1 Create correct weather variable names for the file
weather_var_dict= {"Tn":"tasmin", "Tg":"tas",
          "Tx":"tasmax"
          }
# 2.2 Desired order of column for the output weather file
weather_col_order = ['day','month','Year',  'tas',  'Lat','Lon'] # ['day','month','Year', 'tasmax', 'tas', 'tasmin', 'Lat','Lon']
grid_weather_col_order = ['day','month','Year',  'tg', 'Lat','Lon'] # ['day','month','Year', 'tn', 'tg', 'tx', "rr", 'Lat','Lon']
# 2.3 Observed phenology datasets from vineyard plots
observed_path = glob.glob(join(phenology_path,"*.xlsx"))[0]
# 2.4 Inquire the list of stations in the data file
observed_data = pd.ExcelFile(observed_path)
plot_col = "Plot"
var_col = "Variety"
#save_path_weather = join(weather_station_path,"summary_weather")
#mkdir(save_path_weather)
# 2.5 Define the path to the Iberian datasets
grid_data_path = join(script_drive, "E_OBS_V24")
var_identifiers = ["rr", "tg", "tn","tx"] # ['tasmax', 'tas', 'tasmin',"pr"]
target_dataset = [] # Collect target datasets
for ncfile in glob.glob(join(grid_data_path,"*.nc")):
    if any(varname+"_" in ncfile for varname in var_identifiers):
        target_dataset.append(ncfile)
mkdir(gridded_climate_path)
# 2.6 Iterate over each station
for index, stage in enumerate(observed_data.sheet_names):
    # 2.6.1 Extract the plot specific phenology data into separate .csv file
    observed_stage_data = pd.read_excel(observed_path,sheet_name=stage)
    plots = observed_stage_data[plot_col]
    for plot in plots:
        if index==0: # Since both stages share the same lon, lat, we only need to extract the climate data once
            plot_lon = observed_stage_data.loc[observed_stage_data[plot_col]==plot,"Lon"]
            plot_lat = observed_stage_data.loc[observed_stage_data[plot_col]==plot,"Lat"]
            plot_lon = round(float(plot_lon),3)
            plot_lat = round(float(plot_lat),3)
            # Extract gridded climate datasets for each plot based on their lon and lat
            target_data_list =[]
            for grid_data in target_dataset:
                time_period = xr.open_dataset(grid_data,mask_and_scale=True).time.data # Extract the full time period
                var_shortname = [var for var in var_identifiers if var+"_" in grid_data][0] # Extract the variable short name
                data_series = extract_site_data_series(grid_data, var_shortname, plot_lon,
                                                       plot_lat, time_period) # Extract the time series of data for a given site
                target_data_list.append(data_series) # Append to an existing empty list
            # Concatenate all list datasets 
            merged_df = pd.concat(target_data_list,axis=1, join='inner', ignore_index=False)
            # Create a df with all desired columns
            merged_df = merged_df.assign(day=lambda x:x.index.day, month=lambda x:x.index.month, Year= lambda x:x.index.year,
                             Lon=plot_lon, Lat=plot_lat)
            # Reorder the columns
            merged_df_final = merged_df.reindex(columns = grid_weather_col_order, copy=True)
            # Check for missing values
            if merged_df_final.isnull().values.any():
                merged_df_final.fillna(method="ffill",inplace=True)
            # Export the target df into csv file
            merged_df_final.to_csv(join(gridded_climate_path,"{}_gridded_data.csv".format(plot)), sep=",",header=True,
                                    index=False,encoding ="utf-8")
        # Extract the phenology dataset
        phenology_plot = observed_stage_data.loc[observed_stage_data[plot_col]==plot,~observed_stage_data.columns.isin([plot_col,var_col,"Lat","Lon"])]
        phenology_plot.rename(index={phenology_plot.index.values[0]:"ob_plot{}".format(plot)}, inplace=True)
        # Transpose the df into a single column df
        phenology_plot = phenology_plot.T
        # phenology_plot.rename(columns={0:"phenology_ob"},inplace=True)
        # Check the variety
        variety_name= observed_stage_data.loc[observed_stage_data[plot_col]==plot,var_col].values[0]
        data_save_path = join(phenology_path,variety_name,stage)
        mkdir(data_save_path)
        phenology_plot.to_csv(join(data_save_path, "ob_plot{}".format(plot) +".csv"),
            index=True, header=True,index_label="Year")